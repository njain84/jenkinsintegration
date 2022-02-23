import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from selenium.webdriver.support.ui import Select
import unittest
import openpyxl
from configparser import ConfigParser
import pytest
from selenium.common.exceptions import TimeoutException, WebDriverException


class H323(unittest.TestCase):

    def test_H323_Registration(self):

        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')

        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        mylogin.doLoginadmin()
        time.sleep(5)  # clicking on configuration tab
        self.driver.find_elements_by_class_name('btn')[0].click()
        time.sleep(2)  # selecting signalling setting
        self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
        time.sleep(4)
        checkbox_value = self.driver.find_element_by_xpath("//*[@id='modalForm:isRegToGKEnableCheck_signal1']").get_attribute('checked')
        print('Checkbox is:', checkbox_value)
        if checkbox_value == 'true':
            #self.driver.find_element_by_xpath("//*[@id='modalForm:isRegToGKEnableCheck_signal1']").click()  # this click the Register to  Server
            # time.sleep(3)
            # self.driver.find_element_by_xpath("//*[@id='modalForm:primaryGKIP_signal1']").send_keys('10.97.58.167') # it will put the DMA IP
            # time.sleep(3)    # fill the System Prefix | E.164
            # self.driver.find_element_by_xpath("//*[@id='modalForm:sysPrefix_signal1']").send_keys(parser.get('bug_tracker', 'msh323address'))
            # time.sleep(3)  # fill the System H.323 Alias
            # self.driver.find_element_by_xpath("//*[@id='modalForm:sysH323Alias_signal1']").send_keys(parser.get('bug_tracker', 'msh323address'))

            time.sleep(4)
            self.driver.find_element_by_id('Sip-blt').click()  # click OK
            try:
                time.sleep(3)
                self.driver.find_element_by_id('dialogFirstBt').click()  # OK pop up
                time.sleep(40)
                self.driver.find_element_by_id('dialogFirstBt').click()  # OK another popup
            except WebDriverException as e:
                print('Result:', e)
            time.sleep(5)  # clicking on configuration tab
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(2)  # selecting signalling setting
            self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
            time.sleep(7)  # Now it will see the status using below
            test1_var = self.driver.find_element_by_xpath("//*[@id='isEnablePrimaryGK_signal1']").text
            print(test1_var)
            with open('../output.txt', 'a+') as f:
                f.write("\n\n")
                f.write(test1_var)
            logging.basicConfig(filename='../output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('../output.txt', 'a') as f:
                f.write("\n")
            if test1_var == 'Online':
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1608-Test case is passed')
                    book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                    sheet = book['sample']
                    print('sheet is', sheet)
                    sheet.cell(row=11, column=3).value = 'Pass'
                    print('value is', sheet.cell(row=11, column=3).value)
                    book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1608-Test case is failed')
                    print("EXTMSEN-1608-Test case is failed")
                    time.sleep(5)
        else:
            self.driver.find_element_by_xpath("//*[@id='modalForm:isRegToGKEnableCheck_signal1']").click()  # this click the Register to  Server
            # time.sleep(3)
            # self.driver.find_element_by_xpath("//*[@id='modalForm:primaryGKIP_signal1']").send_keys('10.97.58.167') # it will put the DMA IP
            # time.sleep(3)    # fill the System Prefix | E.164
            # self.driver.find_element_by_xpath("//*[@id='modalForm:sysPrefix_signal1']").send_keys(parser.get('bug_tracker', 'msh323address'))
            # time.sleep(3)  # fill the System H.323 Alias
            # self.driver.find_element_by_xpath("//*[@id='modalForm:sysH323Alias_signal1']").send_keys(parser.get('bug_tracker', 'msh323address'))

            time.sleep(4)
            self.driver.find_element_by_id('Sip-blt').click()  # click OK
            try:
                time.sleep(3)
                self.driver.find_element_by_id('dialogFirstBt').click()  # OK pop up
                time.sleep(44)
                self.driver.find_element_by_id('dialogFirstBt').click()  # OK another popup
            except WebDriverException as e:
                print('Result:', e)
            time.sleep(5)  # clicking on configuration tab
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(2)  # selecting signalling setting
            self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
            time.sleep(7)  # Now it will see the status using below
            test1_var = self.driver.find_element_by_xpath("//*[@id='isEnablePrimaryGK_signal1']").text
            print(test1_var)
            with open('../output.txt', 'a+') as f:
                f.write("\n\n")
                f.write(test1_var)
            logging.basicConfig(filename='../output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('../output.txt', 'a') as f:
                f.write("\n")
            if test1_var == 'Online':
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1608-Test case is passed')
                    book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                    sheet = book['sample']
                    print('sheet is', sheet)
                    sheet.cell(row=11, column=3).value = 'Pass'
                    print('value is', sheet.cell(row=11, column=3).value)
                    book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1608-Test case is failed')
                    print("EXTMSEN-1608-Test case is failed")
                    time.sleep(5)

        self.assertEqual(test1_var, 'Online',"EXTMSEN-1608-Assert error: Test case is failed because status is not Online!")

    if __name__ == '__main__':
        unittest.main()
