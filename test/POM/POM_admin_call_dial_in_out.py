import unittest
import webbrowser

import openpyxl
import pyautogui
import requests
import tkinter as tk
import time
import logging
from openpyxl import Workbook

from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from datetime import datetime
from Login_config.LoginHandler import Loginms
from Login_config.LoginHandleradminportal import Loginmsadmin

from configparser import ConfigParser

class admin_call():

    def __init__(self, driver):

        self.driver = driver

    def admin_dialouttorecord(self):
        #options = webdriver.ChromeOptions()
        #options.add_argument('--headless')
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')#options=options
        myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        #login the admin portal
        myloginadmin.doLoginadmin()
        time.sleep(10)
        # it is extra code to disconnect the call
        try:
            if self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text == 'SIP' or 'H323':
                self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
                time.sleep(4)
                self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        except:
            print('no call')
            # end  the extra code
        time.sleep(9)
        self.driver.find_elements_by_class_name('btn')[2].click()  # opening the Dail out to Record page
        time.sleep(8)
        select = Select(self.driver.find_element_by_id('modalForm:j_idt8'))  # Selecting the drop down
        time.sleep(3)
        select.select_by_visible_text('SIP')  # Selecting the SIP
        time.sleep(5)
        self.driver.find_elements_by_name('modalForm:address_p1')[0].send_keys(parser.get('bug_tracker', 'EPIP'))
        time.sleep(5)
        self.driver.find_element_by_xpath("//button[@id='confirmDialOut']").click()  # Dail the call
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        time.sleep(5)  # click on call tab
        self.driver.find_elements_by_class_name('panel')[1].click()
        # New code to enable the proper call info on UI , it is extra we added for EXTMSEN-1646
        time.sleep(3)
        self.driver.find_element_by_id('a_1000').click()
        time.sleep(1)
        self.driver.find_element_by_id('a_1001').click()
        # This extra code is ended here
        time.sleep(60)
        # click on page to get VOIP protocol info
        test1_var = self.driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a+') as f:
            f.write("\n\n")
            f.write(test1_var)
        time.sleep(3)
        # click on Details, Pop up will reflect
        self.driver.find_element_by_id('portalPopoverCall-0').click()
        time.sleep(2)
        test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        time.sleep(3)
        with open('output.txt', 'a') as f:
            f.write("\n")
            f.write(test_var)
        pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
        print("VOIP protocol is :", pertext)
        # this live_streaming_tab_info is to get the info od this tab
        time.sleep(1)
        live_streaming_tab_info = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[9]").text
        
        # This code is to fetch the information from detail tab and convert it into list. at last it will split the list and fetch the call rate.
        x = [test_var]
        print('Total values :',test_var)
        call_rate = x[0].split('\n')
        print('Call_rate :',call_rate)
        call_rate1 = call_rate[5]  
        call_rate2 = call_rate1.split('Video Call Rate (Kbps) ',1)
        final_call_rate = call_rate2[1]
        # -----------------------------------------------------
        logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                            format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open('output.txt', 'a') as f:
            f.write("\n")
        if pertext == 'SIP':
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write('EXTMSEN-Test case is passed')
                f.write("\n")
        time.sleep(3)
        self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        time.sleep(5)
        print(pertext)
        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        time.sleep(2)
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('time is', createdtime)
        #To get the file name, we are adding another x path below
        file_name = self.driver.find_element_by_xpath('//*[@id="archive-table-row-0"]/td[2]').text
        print(file_name)
        require_file_name = file_name.startswith('Autotest')
        print('Require_file_name',require_file_name)
        time.sleep(2)
        #it will go select the file and click on transcoding to check the high profile of live streaming
        time.sleep(1)
        self.driver.find_element_by_name('mainForm:j_idt26:0:j_idt28').click()
        time.sleep(1)
        self.driver.find_elements_by_class_name('btn')[5].click()
        time.sleep(1)
        live_mp4_profile = self.driver.find_element_by_xpath('//*[@id="archiveFile-table-row-1"]/td[12]').text
        print('live_profile:',live_mp4_profile)
        
        if createdtime >= dt_string and pertext == 'SIP':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=5, column=3).value = 'Pass'
            print('value is', sheet.cell(row=5, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print("failed")
        # it will return below mentioned variables 
        return [createdtime,dt_string,pertext,createdtime,final_call_rate,live_streaming_tab_info,require_file_name,live_mp4_profile]

    
# for Dial In
    def admin_dialin(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe',options=options)  #options=options
        myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        #login the admin portal
        myloginadmin.doLoginadmin()
        time.sleep(5)  # click on call tab
        self.driver.find_elements_by_class_name('panel')[1].click()
        time.sleep(60)
        # click on page to get VOIP protocol info
        test1_var = self.driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a+') as f:
            f.write("\n\n")
            f.write(test1_var)
        time.sleep(5)
        # click on Details, Pop up will reflect
        self.driver.find_element_by_id('portalPopoverCall-0').click()
        test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        with open('output.txt', 'a') as f:
            f.write("\n")
            f.write(test_var)
        pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
        print("VOIP protocol is :", pertext)
        logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                            format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open('output.txt', 'a') as f:
            f.write("\n")
        if pertext == 'SIP':
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write('EXTMSEN-1625-Test case is passed')
                f.write("\n")
        time.sleep(3)
        self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        time.sleep(5)
        print(pertext)
        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('file creation time is', createdtime)
        time.sleep(2)
        if createdtime >= dt_string and pertext == 'SIP':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=5, column=3).value = 'Pass'
            print('value is', sheet.cell(row=5, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print("failed")
        # It will return below mentioned parameters
        return [createdtime,pertext]

# H323 Dial out:

    def admin_dialouttorecord_H323(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome(
            'C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe',options=options)  # options=options
        myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        # login the admin portal
        myloginadmin.doLoginadmin()
        time.sleep(10)
        # it is extra code to disconnect the call
        try:
            if self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text == 'SIP' or 'H323':
                self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
                time.sleep(4)
                self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        except:
            print('no call')
            # end  the extra code
        time.sleep(9)
        self.driver.find_elements_by_class_name('btn')[2].click()  # opening the Dial out to Record page
        time.sleep(8)

        select = Select(self.driver.find_element_by_id('modalForm:j_idt8'))  # Selecting the drop down
        time.sleep(3)
        select.select_by_visible_text('H.323')  # Selecting the H323
        time.sleep(5)
        self.driver.find_elements_by_name('modalForm:address_p1')[0].send_keys(parser.get('bug_tracker', 'EPIP_H323'))
        time.sleep(5)
        self.driver.find_element_by_xpath("//button[@id='confirmDialOut']").click()  # Dail the call
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        time.sleep(5)  # click on call tab
        self.driver.find_elements_by_class_name('panel')[1].click()
        time.sleep(60)
        # click on page to get VOIP protocol info
        test1_var = self.driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a+') as f:
            f.write("\n\n")
            f.write(test1_var)
        time.sleep(3)
        # click on Details, Pop up will reflect
        self.driver.find_element_by_id('portalPopoverCall-0').click()
        time.sleep(2)
        test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        time.sleep(3)
        with open('output.txt', 'a') as f:
            f.write("\n")
            f.write(test_var)
        pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
        print("VOIP protocol is :", pertext)
        # This code is to fetch the information from detail tab and convert it into list. at last it will split the list and fetch the call rate.
        x = [test_var]
        print('Total values :',test_var)
        call_rate = x[0].split('\n')
        print('Call_rate :',call_rate)
        call_rate1 = call_rate[5]
        call_rate2 = call_rate1.split('Video Call Rate (Kbps) ',1)
        final_call_rate = call_rate2[1]
        # -----------------------------------------------------
        logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                            format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open('output.txt', 'a') as f:
            f.write("\n")
        if pertext == 'H323':
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write('EXTMSEN-1625-Test case is passed')
                f.write("\n")
        time.sleep(3)
        self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        time.sleep(5)
        print(pertext)
        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('time is', createdtime)
        time.sleep(2)
        if createdtime >= dt_string and pertext == 'H323':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=5, column=3).value = 'Pass'
            print('value is', sheet.cell(row=5, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print("failed")
        self.driver.quit()
        return [createdtime, dt_string, pertext,final_call_rate]
    
    def admin_dialouttorecord_choose_VRR(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        print("Hello....")
        parser = ConfigParser()
        configFilePath = r'C:\PycharmProjects\PycharmProjects\Media_Suite_Smoke\test\config.ini'
        parser.read(configFilePath)
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe',options=options) #options=options
        myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        # login the admin portal
        myloginadmin.doLoginadmin()
        #login the admin portal
        time.sleep(10)
        # it is extra code to disconnect the call
        try:
            if self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text == 'SIP' or 'H323':
                self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
                time.sleep(4)
                self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        except:
            print('no call')
            # end  the extra code
        time.sleep(9)
        self.driver.find_elements_by_class_name('btn')[2].click()  # opening the Dail out to Record page
        time.sleep(8)
        select = Select(self.driver.find_element_by_id('modalForm:j_idt8'))  # Selecting the drop down
        time.sleep(3)
        select.select_by_visible_text('SIP')  # Selecting the SIP
        # it will select the Vrr
        select=Select(self.driver.find_element_by_id('modalForm:vrrName'))
        select.select_by_visible_text('test')
        # it will enter the EPIp
        time.sleep(5)
        self.driver.find_elements_by_name('modalForm:address_p1')[0].send_keys(parser.get('bug_tracker', 'EPIP'))
        time.sleep(5)
        self.driver.find_element_by_xpath("//button[@id='confirmDialOut']").click()  # Dail the call
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        time.sleep(5)  # click on call tab
        self.driver.find_elements_by_class_name('panel')[1].click()
        time.sleep(60)
        # click on page to get VOIP protocol info
        test1_var = self.driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a+') as f:
            f.write("\n\n")
            f.write(test1_var)
        time.sleep(3)
        # click on Details, Pop up will reflect
        self.driver.find_element_by_id('portalPopoverCall-0').click()
        test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        time.sleep(3)
        self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        self.driver.implicitly_wait(2)
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('time is', createdtime)
        time.sleep(2)
        return [createdtime,dt_string]


