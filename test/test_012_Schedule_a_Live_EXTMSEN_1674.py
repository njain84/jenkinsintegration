import time
import logging
from selenium import webdriver
import pyautogui
from Login_config.LoginHandleradminportal import Loginmsadmin
from configparser import ConfigParser
import unittest
import openpyxl
import allure
import pytest
from configparser import ConfigParser
from Login_config.LoginHandler import Loginms
from datetime import datetime
from datetime import datetime as dt


class Schedulelive(unittest.TestCase):

    def test_schedule_live(self):
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginms('admin', 123, parser.get('bug_tracker', 'urluserportal'), self.driver)
        myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        mylogin.doLogin()
        # click on Administrator
        time.sleep(10)
        self.driver.find_elements_by_class_name("avatar_active")[0].click()
        # click on My media center
        time.sleep(3)
        self.driver.find_elements_by_class_name("btn")[2].click()
        # click on Events
        time.sleep(3)
        #self.driver.find_element_by_xpath("// *[ @ id = 'myTab'] / li[2] / a / span[1]").click()
        self.driver.find_element_by_link_text('Events').click()
        time.sleep(3)
        # click on Schedule live
        self.driver.find_element_by_xpath("// *[ @ id = 'MyLiveEvent'] / div / h2 / span[2] / a[2]").click()
        time.sleep(2)
        # fill title
        self.driver.find_element_by_xpath("// *[ @ id = 'eventForm'] / div[1] / div / div[1] / div / input").send_keys('Smoke')
        # fill Meeting room
        time.sleep(2)
        self.driver.find_element_by_xpath("// *[ @ id = 'eventForm'] / div[1] / div / div[3] / div / input").send_keys(parser.get('bug_tracker', 'vmr'))
        # it will click save
        time.sleep(2)
        self.driver.find_element_by_xpath("// *[ @ id = 'eventForm'] / div[5] / div[2] / div / a[2]").click()
        time.sleep(2)
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        # it will print start and end time
        time.sleep(3)
        print(self.driver.find_element_by_xpath("// *[ @ id = 'MyLiveEvent'] / div / div[3] / div / div[4] / div / div[2] / div / div / div[1] / i"))
        time.sleep(3)
        # it will tell start and end time
        print('time is', self.driver.find_element_by_xpath("//*[@id='MyLiveEvent']/div/div[3]/div/div[4]/div/div[2]/div/div/div[1]").text)
        time.sleep(3)
        print('Date is', self.driver.find_element_by_xpath("// *[ @ id = 'MyLiveEvent'] / div / div[3] / div / div[4] / div / div[1] / small").text)
        minutes = dt.now().minute
        print(minutes)
        i = 1
        while i < 17:
            print(i)
            minutes = dt.now().minute
            if minutes == 15 or minutes == 30 or minutes == 45 or minutes == 00:
                break
            time.sleep(60)
            i += 1
        time.sleep(3)
        self.driver.refresh()
        # click on Administrator
        time.sleep(10)
        self.driver.find_elements_by_class_name("avatar_active")[0].click()
        # click on My media center
        time.sleep(3)
        self.driver.find_elements_by_class_name("btn")[2].click()
        # click on Events
        time.sleep(3)
        self.driver.find_element_by_xpath("// *[ @ id = 'myTab'] / li[2] / a / span[1]").click()
        # It is providing  the value of class(show/ng hide)
        time.sleep(10)
        now_image = self.driver.find_element_by_xpath("//*[@id='MyLiveEvent']/div/div[3]/div/div[4]/div[1]/div[2]/div/div/div[1]/i").get_attribute('class')
        print('now is', now_image)

        #print(self.driver.find_element_by_xpath("//*[@id='MyLiveEvent']/div/div[3]/div/div[4]/div[1]/div[2]/div/div/div[1]/i").get_attribute('ng-show'))
        #print(self.driver.find_element_by_xpath("//*[@id='MyLiveEvent']/div/div[3]/div/div[4]/div[1]/div[2]/div/div/div[1]/i").value_of_css_property('background-color'))
        # It will start the call
        time.sleep(3)
        self.driver.find_element_by_xpath("// *[ @ id = 'MyLiveEvent'] / div / div[3] / div / div[4] / div / div[2] / div / div / div[4] / button / a").click()
        time.sleep(5)
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)

        time.sleep(30)
        # it will hangup the call
        self.driver.find_element_by_xpath("//*[@id='menuContent']/div/div/h2/div[2]/a").click()
        time.sleep(3)
        self.driver.find_element_by_xpath("// *[ @ id = 'dialog'] / div / div / div[3] / div[2] / button").click()
        time.sleep(5)
        myloginadmin.doLoginadmin()

        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        time.sleep(3)
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('File creation time is', createdtime)
        if createdtime >= dt_string and now_image == 'now':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=13, column=3).value = 'Pass'
            print('value is', sheet.cell(row=13, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print("failed")

        assert createdtime >= dt_string, "EXTMSEN-1674-Assert error: file is not created"
        self.assertEqual(now_image, 'now',"EXTMSEN-1674-Assert error: Test case is failed because status is not now!")

        time.sleep(10)



