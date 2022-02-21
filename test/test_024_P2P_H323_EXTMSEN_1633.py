import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from Login_config.LoginHandler import Loginms
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from .EPlogin_call import Eplogin
from .EPlogin_callhangup import Ephangup
from datetime import datetime
from configparser import ConfigParser
import openpyxl
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
from .POM.POM_Registration import admin_Registration

class DialtoepH323(unittest.TestCase):

    def test_dial_out_H323_EP(self):

        try:
            print("Hello....")
            parser = ConfigParser()
            parser.read('config.ini')
            self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
            # it will uregoster the MS from DMA first
            unregister = admin_Registration(self.driver)
            actualunregister = unregister.test_H323_unregister()
            print(actualunregister)
            time.sleep(4)
            mylogin = Loginms('admin', 123, parser.get('bug_tracker', 'urluserportal'), self.driver)
            myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
            mylogin.doLogin()
            print("Hi")
            time.sleep(15)  # opening the recording page to dail
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(5)
            # click P2P
            time.sleep(2)
            self.driver.find_element_by_xpath('''//button[@ng-if="supportedSignalingType != 'sip' && hasLiveStreamingPorts"]''').click()
            time.sleep(5)
            # Fill the title
            self.driver.find_element_by_name('subject').send_keys('test',Keys.TAB,parser.get('bug_tracker', 'H323_1633'),Keys.TAB,parser.get('bug_tracker', 'EPH323-1'))
            time.sleep(5)
            self.driver.find_element_by_id('dialBtnInSR').click()  # Dailing the call
            now = datetime.now()
            dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
            print('Current time', dt_string)
            time.sleep(10)
            myloginadmin.doLoginadmin()
            time.sleep(40)
            pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text  # it will check the protocol
            print("VOIP protocol is :", pertext)
            time.sleep(5)
            logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('output.txt', 'a') as f:
                f.write("\n")
            try:
                assert pertext == "H323", "EXTMSEN-1628-Assert error: Test case is failed because protocol is not H323!"
            except AssertionError as e:
                print('Result:', e)
                logging.error(str(e))
            if pertext == 'H323':
                with open('output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1619-Test case is passed')
                    f.write("\n")
            self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
            time.sleep(4)
            self.driver.find_element_by_xpath(
                "//*[@id='dialogFirstBt']").click()  # Click OK after hangup at admin portal
            time.sleep(70)  # Now it will go to Archive and check the file
            self.driver.find_element_by_id('a_1002').click()
            time.sleep(3)
            self.driver.find_element_by_id('a_1003').click()
            time.sleep(4)
            createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
            print('File creation time is', createdtime)
            time.sleep(3)
            if createdtime >= dt_string and pertext == 'H323':
                print("test case is passed")
                book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                sheet = book['sample']
                print('sheet is', sheet)
                sheet.cell(row=12, column=3).value = 'Pass'
                print('value is', sheet.cell(row=12, column=3).value)
                book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                print("failed")

            assert createdtime >= dt_string, "EXTMSEN-1633-Assert error: file is not created"
            self.assertEqual(pertext, 'H323',"EXTMSEN-1633-Assert error: Test case is failed because protocol is not H323!")
            self.assertEqual(actualunregister, 'Offline',"EXTMSEN-1633-Assert error: Test case is failed because registration still active")
        finally:
            try:
                print('This is always executed')
                driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
                parser = ConfigParser()
                parser.read('config.ini')
                ephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)
                ephangup.doepcallhangup()

            except WebDriverException as e:
                print('Result:', e)

    if __name__ == '__main__':
        unittest.main()
