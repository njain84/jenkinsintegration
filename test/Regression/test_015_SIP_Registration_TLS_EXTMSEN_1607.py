import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
#from selenium.common.exceptions import Exception
from selenium.webdriver.support.ui import Select
import unittest
import openpyxl
from openpyxl import Workbook
from configparser import ConfigParser
import pytest
from selenium.common.exceptions import TimeoutException, WebDriverException

class TLS(unittest.TestCase):

    def test_TLS_registration(self):

        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        mylogin.doLoginadmin()
        time.sleep(5)  # clicking on configuration tab
        self.driver.find_elements_by_class_name('btn')[0].click()
        time.sleep(2)  # selecting signalling setting
        self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
        time.sleep(2)  # selecting sip
        self.driver.find_element_by_xpath("//*[@id='signalTab_signal1']/li[2]/a").click()
        #It will check the check tick mark in registration server check box
        checkbox_value = self.driver.find_element_by_id('modalForm:isRegSIPCheck_signal1').get_attribute('checked')
        print('Checkbox is:',checkbox_value)
        if checkbox_value == 'true':
            
            time.sleep(3)  # will click and select the transport protocol
            self.select = Select(self.driver.find_element_by_id('modalForm:SIP-TransportType_signal1'))
            self.select.select_by_visible_text('TLS')
            time.sleep(3)
            # self.driver.find_element_by_xpath("//*[@id='modalForm:SIPServerAddr_signal1']").send_keys('10.97.58.167') # it will put the DMA IP
            # time.sleep(4)
            # this will fill the user name for registration
            # self.driver.find_element_by_xpath("// *[ @ id = 'modalForm:regUserName_signal1']").send_keys(parser.get('bug_tracker', 'mssipaddress'))
            time.sleep(4)
            self.driver.find_element_by_id('Sip-blt').click()
            try:
                time.sleep(3)
                self.driver.find_element_by_id('dialogFirstBt').click()
                time.sleep(30)
                self.driver.find_elements_by_class_name('btn')[10].click()
            except WebDriverException as e:
                print('Result:', e)
            time.sleep(5)  # clicking on configuration tab
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(2)  # selecting signalling setting
            self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
            time.sleep(2)  # selecting sip
            self.driver.find_element_by_xpath("//*[@id='signalTab_signal1']/li[2]/a").click()
            time.sleep(7)  # Now it will see the status using below
            test1_var = self.driver.find_element_by_xpath("//*[@id='SIPRegStatus_signal1']").text
            protocol = self.driver.find_element_by_xpath("//*[@id='modalForm:SIP-TransportType_signal1']").text
            print("sip server is:", test1_var)
            print("protocol is:", protocol)
            protocol_value = self.driver.find_element_by_id('modalForm:SIP-TransportType_signal1').get_attribute("value")
            print('transport protocol is:', protocol_value)
            with open('../output.txt', 'a+') as f:
                f.write("\n\n")
                f.write(test1_var)
            logging.basicConfig(filename='../output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('../output.txt', 'a') as f:
                f.write("\n")
            if test1_var == 'Online':
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1606-Test case is passed')
                    f.write("\n")
                    book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                    sheet = book['sample']
                    print('sheet is', sheet)
                    sheet.cell(row=3, column=3).value = 'Pass'
                    print('value is', sheet.cell(row=3, column=3).value)
                    book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')

            else:
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1606-Test case is failed')
                    print("EXTMSEN-1606-Test case is failed")

                    time.sleep(5)
        else:
            print('Executing else part')

            self.driver.find_element_by_id('modalForm:isRegSIPCheck_signal1').click()  #this click the Register to SIP Server
            time.sleep(3)  # will click and select the transport protocol
            self.select = Select(self.driver.find_element_by_id('modalForm:SIP-TransportType_signal1'))
            self.select.select_by_visible_text('TLS')
            time.sleep(3)
            # self.driver.find_element_by_xpath("//*[@id='modalForm:SIPServerAddr_signal1']").send_keys('10.97.58.167') # it will put the DMA IP
            # time.sleep(4)
            # this will fill the user name for registration
            # self.driver.find_element_by_xpath("// *[ @ id = 'modalForm:regUserName_signal1']").send_keys(parser.get('bug_tracker', 'mssipaddress'))
            time.sleep(4)
            self.driver.find_element_by_id('Sip-blt').click()
            try:
                time.sleep(3)
                self.driver.find_element_by_id('dialogFirstBt').click()
                time.sleep(30)
                self.driver.find_elements_by_class_name('btn')[10].click()
            except WebDriverException as e:
                print('Result:', e)
            time.sleep(5)  # clicking on configuration tab
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(2)  # selecting signalling setting
            self.driver.find_element_by_id('headerForm:j_idt10:1:j_idt12').click()
            time.sleep(2)  # selecting sip
            self.driver.find_element_by_xpath("//*[@id='signalTab_signal1']/li[2]/a").click()
            time.sleep(7)  # Now it will see the status using below
            test1_var = self.driver.find_element_by_xpath("//*[@id='SIPRegStatus_signal1']").text
            protocol = self.driver.find_element_by_xpath("//*[@id='modalForm:SIP-TransportType_signal1']").text
            print("sip server is:", test1_var)
            print("protocol is:", protocol)
            protocol_value = self.driver.find_element_by_id('modalForm:SIP-TransportType_signal1').get_attribute("value")
            print('transport protocol is:', protocol_value)
            
           
            with open('../output.txt', 'a+') as f:
                f.write("\n\n")
                f.write(test1_var)
            logging.basicConfig(filename='../output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('../output.txt', 'a') as f:
                f.write("\n")
            if test1_var == 'Online':
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1606-Test case is passed')
                    f.write("\n")
                    book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                    sheet = book['sample']
                    print('sheet is', sheet)
                    sheet.cell(row=3, column=3).value = 'Pass'
                    print('value is', sheet.cell(row=3, column=3).value)
                    book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')

            else:
                with open('../output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1606-Test case is failed')
                    print("EXTMSEN-1606-Test case is failed")

                    time.sleep(5)
        self.assertEqual(test1_var, 'Online',"EXTMSEN-1607-Assert error: Test case is failed because status is not Online!")
        assert protocol_value == '2', "EXTMSEN-1606-Protocol is not TLS"
        self.driver.quit()

    if __name__ == '__main__':
        unittest.main()





