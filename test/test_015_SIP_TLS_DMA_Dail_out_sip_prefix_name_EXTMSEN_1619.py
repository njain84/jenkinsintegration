import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from Login_config.LoginHandler import Loginms
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from .EPlogin_call import Eplogin
from .EPlogin_callhangup import Ephangup
from datetime import datetime
from configparser import ConfigParser
import openpyxl
from selenium.common.exceptions import TimeoutException, WebDriverException

class Dialtoepsip_prefixname(unittest.TestCase):

    def test_dial_out_SIP_EP_Prefixname(self):
        try:
            print("Hello....")
            parser = ConfigParser()
            parser.read('config.ini')
            self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
            mylogin = Loginms('admin', 123, parser.get('bug_tracker', 'urluserportal'), self.driver)
            myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
            mylogin.doLogin()
            print("Hi")
            time.sleep(15)  # opening the recording page to dail
            self.driver.find_elements_by_class_name('btn')[0].click()
            time.sleep(5)
            self.driver.find_element_by_id('meetingNumber').send_keys(parser.get('bug_tracker', 'rpdsipaddress'))
            time.sleep(5)
            # Fill the title
            self.driver.find_elements_by_name('subject')[0].send_keys('test')
            time.sleep(5)
            # will click on advances option
            self.driver.find_element_by_xpath(
                "//*[@id='firstModal']/div/div/div[2]/div/div/div/form/div[3]/div[1]/div/small/a").click()
            time.sleep(2)
            self.driver.find_element_by_xpath(
                "//*[@id='firstModal']/div/div/div[2]/div/div/div/form/div[3]/div[2]/div/div[2]/div/label[1]/input").click()
            time.sleep(5)
            self.driver.find_element_by_id('dialBtnInSR').click()  # Dailing the call
            now = datetime.now()
            dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
            print('Current time', dt_string)
            time.sleep(10)
            myloginadmin.doLoginadmin()
            time.sleep(40)
            pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text  # it will check the protocol
            print("VOIP protocol is :", pertext)
            time.sleep(5)
            logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('output.txt', 'a') as f:
                f.write("\n")
            try:
                assert pertext == "SIP", "EXTMSEN-1628-Assert error: Test case is failed because protocol is not SIP!"
            except AssertionError as e:
                print('Result:', e)
                logging.error(str(e))
            if pertext == 'SIP':
                with open('output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1619-Test case is passed')
                    f.write("\n")
            self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
            time.sleep(4)
            self.driver.find_element_by_xpath(
                "//*[@id='dialogFirstBt']").click()  # Click OK after hangup at admin portal
            time.sleep(70)  # Now it will go to Archive and check the file
            self.driver.find_element_by_id('a_1002').click()
            time.sleep(3)
            self.driver.find_element_by_id('a_1003').click()
            time.sleep(4)
            createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
            print('File creation time is', createdtime)
            time.sleep(3)
            if createdtime >= dt_string and pertext == 'SIP':
                print("test case is passed")
                book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                sheet = book['sample']
                print('sheet is', sheet)
                sheet.cell(row=12, column=3).value = 'Pass'
                print('value is', sheet.cell(row=12, column=3).value)
                book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                print("failed")

            assert createdtime >= dt_string, "EXTMSEN-1619-Assert error: file is not created"
            self.assertEqual(pertext, 'SIP',"EXTMSEN-1619-Assert error: Test case is failed because protocol is not SIP!")
        finally:
            try:
                print('This is always executed')
                driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
                parser = ConfigParser()
                parser.read('config.ini')
                ephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)
                ephangup.doepcallhangup()

            except WebDriverException as e:
                print('Result:', e)

    if __name__ == '__main__':
        unittest.main()
