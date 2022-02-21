import xlsxwriter
import openpyxl
from openpyxl import Workbook


def create_excel():
    print('start')
    book = openpyxl.load_workbook('C:\\Users\\njain\\Desktop\\demo.xlsx')
    sheet = book.get_sheet_by_name('demo')
    sheet.cell(row=5, column=1).value = 'Appended Data'

    def main():
        book = openpyxl.load_workbook('C:\\Users\\njain\\Desktop\\demo.xlsx')
        sheet = book.get_sheet_by_name('demo')
        sheet.cell(row=5, column=1).value = 'Appended Data'
        book.save('demo.xlsx')

    #c1.value = ['Ankit']




# Create the excel
    #wb = openpyxl.Workbook()
    #sheet = wb.active
    #c1 = sheet.cell(row=2, column=2)
    #c1.value = "ANKIT"
    #wb.save("C:\\Users\\njain\\Desktop\\demo.xlsx")
    sheet.append(c1)



create_excel()



