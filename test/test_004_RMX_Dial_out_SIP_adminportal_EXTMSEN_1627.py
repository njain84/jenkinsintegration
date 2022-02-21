import webbrowser
import requests
import tkinter as tk
import time
import logging
import sys
import os
from selenium import webdriver
from .EPlogin_call import Eplogin
#from Login_config.EPlogin_call import Eplogin
#from Login_config.EPlogin_callhangup import Ephangup
from .EPlogin_call import Eplogin
from .EPlogin_callhangup import Ephangup
from Login_config.LoginHandleradminportal import Loginmsadmin
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from selenium.common.exceptions import TimeoutException, WebDriverException
from datetime import datetime
from configparser import ConfigParser
import openpyxl


class Dialtoephsipvmr(unittest.TestCase):

    def test_sip_vmr_call(self):
        try:
            print("Hello....")
            parser = ConfigParser()
            parser.read('config.ini')
            self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
            mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
            mylogin.doLoginadmin()
            time.sleep(10)
            # it is extra code to disconnect the call
            try:
                if self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text == 'SIP' or 'H323':
                    self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
                    time.sleep(4)
                    self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
            except:
                print('no call')
            time.sleep(10)
            self.driver.find_elements_by_class_name('btn')[2].click()  # opening the Dail out to Record page
            time.sleep(5)
            self.driver.find_elements_by_name('modalForm:address_p1')[0].send_keys(parser.get('bug_tracker', 'vmr'))
            # time.sleep(5)
            # driver.find_elements_by_id('modalForm:j_idt8')
            time.sleep(10)
            self.driver.find_element_by_xpath("//button[@id='confirmDialOut']").click()  # Dail the call and OK
            now = datetime.now()
            dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
            print('Current time', dt_string)
            time.sleep(10)
            try:
                myepcall = Eplogin(parser.get('bug_tracker', 'epurl'), self.driver)  # Login the EP
                myepcall.doepLogin()  # EP will dial the conference
            except Exception as e:
                print('EP login error:', e)
                mylogin.doLoginadmin()  # login the admin portal
                time.sleep(5)  # click on call tab
                self.driver.find_elements_by_class_name('panel')[1].click()
                time.sleep(3)
                self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call

            time.sleep(5)
            mylogin.doLoginadmin()  # login the admin portal
            time.sleep(5)  # click on call tab
            self.driver.find_elements_by_class_name('panel')[1].click()
            time.sleep(60)  # click on page to get VOIP protocol info
            # driver.find_element_by_class_name('table')[0].click()
            # time.sleep(2)
            test1_var = self.driver.find_elements_by_class_name('table')[0].text
            with open('output.txt', 'a') as f1:
                f1.write("\n")
                f1.write(test1_var)
            time.sleep(3)  # click on Details, Pop will reflect
            self.driver.find_element_by_id('portalPopoverCall-0').click()
            test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write(test_var)
            pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
            print("VOIP protocol is :", pertext)
            logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('output.txt', 'a') as f:
                f.write("\n")
            try:
                assert pertext == "SIP", "EXTMSEN-1627-Assert error: Test case is failed because protocol is not SIP!"
            except AssertionError as e:
                print('Result:', e)
                logging.error(str(e))
            if pertext == 'SIP':
                with open('output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1627-Test case is passed')
                    f.write("\n")

            time.sleep(3)
            self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
            time.sleep(4)
            self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
            time.sleep(70)  # Now it will go to Archive and check the file
            self.driver.find_element_by_id('a_1002').click()
            self.driver.find_element_by_id('a_1003').click()
            time.sleep(4)
            createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
            print('time is', createdtime)
            time.sleep(5)
            try:
                myephangup = Ephangup(parser.get('bug_tracker', 'epurl'), self.driver)  # login the EP
                myephangup.doepcallhangup()  # EP will hangup the call
            except Exception as e:
                print('EP disconnection error:', e)
            time.sleep(2)
            if createdtime >= dt_string and pertext == 'SIP':
                print("test case is passed")
                book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                sheet = book['sample']
                print('sheet is', sheet)
                sheet.cell(row=6, column=3).value = 'Pass'
                print('value is', sheet.cell(row=6, column=3).value)
                book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                print("failed")

            assert createdtime >= dt_string, "EXTMSEN-1627-Assert error: file is not created"
            self.assertEqual(pertext, 'SIP',
                             "EXTMSEN-1627-Assert error: Test case is failed because protocol is not SIP!")
        finally:
            try:
                print('This is always executed')
                driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
                parser = ConfigParser()
                parser.read('config.ini')
                ephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)
                ephangup.doepcallhangup()

            except WebDriverException as e:
                print('Result:', e)

    if __name__ == '__main__':
        unittest.main()
