import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from datetime import datetime
import openpyxl
from openpyxl import Workbook

class Dialtoeph323(unittest.TestCase):

    def test_h323_p2p_call(self):

        print("Hello....")

        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        mylogin.doLoginadmin()
        time.sleep(10)
        # it is extra code to disconnect the call
        try:
            if self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text == 'SIP' or 'H323':
                self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
                time.sleep(4)
                self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        except:
            print('no call')
            #end  the extra code
        time.sleep(9)
        self.driver.find_elements_by_class_name('btn')[2].click()  # opening the Dail out to Record page
        time.sleep(8)
        select = Select(self.driver.find_element_by_id('modalForm:j_idt8'))  # Selecting the drop down
        time.sleep(3)
        select.select_by_visible_text('H.323')  # Selecting the H323
        time.sleep(5)
        self.driver.find_elements_by_name('modalForm:address_p1')[0].send_keys(parser.get('bug_tracker', 'rpdH323address'))
        time.sleep(5)
        self.driver.find_element_by_xpath("//button[@id='confirmDialOut']").click()  # Dail the call
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time',dt_string)
        time.sleep(5)  # click on call tab
        self.driver.find_elements_by_class_name('panel')[1].click()
        time.sleep(60)  # click on page to get VOIP protocol info
        # driver.find_element_by_class_name('table')[0].click()
        # time.sleep(2)
        test1_var = self.driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a+') as f:
            f.write("\n\n")
            f.write(test1_var)
        time.sleep(3)  # click on Details, Pop will reflect
        self.driver.find_element_by_id('portalPopoverCall-0').click()
        test_var = self.driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        with open('output.txt', 'a') as f:
            f.write("\n")
            f.write(test_var)
        pertext = self.driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
        print("VOIP protocol is :", pertext)
        logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                            format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open('output.txt', 'a') as f:
            f.write("\n")
        try:
            self.assertEqual(pertext, 'H323',"EXTMSEN-1625-Assert error: Test case is failed because protocol is not H323!")
        except AssertionError as e:
            print('Result:', e)
            logging.error(str(e))
        if pertext == 'H323':
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write('EXTMSEN-1625-Test case is passed')
                f.write("\n")

        # with open('output.txt') as f:
        # if 'Detail Call Information' in f.read():
        # print("test cases is Passed")
        # test = "Test cases is Passed"
        # with open('output.txt', 'a') as f:
        # f.write("\n")x
        # f.write(test)

        # else:
        # print("Fail")
        # test = "Test case is Failed"
        # with open('output.txt', 'a') as f:
        # f.write(test)
        time.sleep(3)
        self.driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()
        time.sleep(5)
        print(pertext)
        time.sleep(70)  # Now it will go to Archive and check the file
        self.driver.find_element_by_id('a_1002').click()
        self.driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = self.driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('time is', createdtime)
        time.sleep(2)
        if createdtime >= dt_string and pertext == 'H323':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=5, column=3).value = 'Pass'
            print('value is', sheet.cell(row=5, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print("failed")
        assert createdtime >= dt_string, "EXTMSEN-1625-Assert error: file is not created"
        self.assertEqual(pertext , 'H323',"EXTMSEN-1625-Assert error: Test case is failed because protocol is not H323!")



    if __name__ == '__main__':
        unittest.main()




