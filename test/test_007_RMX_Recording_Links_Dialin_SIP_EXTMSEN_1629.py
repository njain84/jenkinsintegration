import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from Login_config.LoginHandler import Loginms
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from .Login_config.EPlogin_call import Eplogin
from .Login_config.EPlogin_callhangup import Ephangup
from .Login_config.RMX_Manager_SIP import RMXmanagerwithSIPstartrecording
from .Login_config.stop_recording import RMXmanagerwithstoptrecording
from datetime import datetime
import openpyxl
import pyautogui
from configparser import ConfigParser



class Dialoutrecordinglink(unittest.TestCase):

    def test_dialoutfromRMX(self):
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), driver)
        #try:
            #myepcall = Eplogin(parser.get('bug_tracker', 'epurl'), driver)  # Login the EP
            #myepcall.doepLogin()  # EP will dial the conference
        #except Exception as e:
            #print('EP login error:', e)
            #myepcall = Eplogin(parser.get('bug_tracker', 'epurl'), driver)  # Login the EP
            #myepcall.doepLogin()  # EP will dial the conference
        now = datetime.now()
        dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
        print('Current time', dt_string)
        time.sleep(5)

        # Rmx will start the recording through RMX manager
        rmxhSIPrecordinglogin = RMXmanagerwithSIPstartrecording(pyautogui)
        rmxhSIPrecordinglogin.RMXmanagerSIP()
        time.sleep(5)
        mylogin.doLoginadmin()  # login the admin portal
        time.sleep(5)  # click on call tab
        driver.find_elements_by_class_name('panel')[1].click()
        time.sleep(60)  # click on page to get VOIP protocol info
        # driver.find_element_by_class_name('table')[0].click()
        # time.sleep(2)
        test1_var = driver.find_elements_by_class_name('table')[0].text
        with open('output.txt', 'a') as f1:
            f1.write("\n")
            f1.write(test1_var)
        time.sleep(3)  # click on Details, Pop will reflect
        driver.find_element_by_id('portalPopoverCall-0').click()
        test_var = driver.find_elements_by_class_name('popover')[0].text  # it will save the data
        with open('output.txt', 'a') as f:
            f.write("\n")
            f.write(test_var)
        pertext = driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
        print("VOIP protocol is :", pertext)
        logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                            format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open('output.txt', 'a') as f:
            f.write("\n")
        try:
            assert pertext == "SIP", "EXTMSEN-1629-Assert error: Test case is failed because protocol is not H323!"
        except AssertionError as e:
            print('Result:', e)
            logging.error(str(e))
        if pertext == 'H323':
            with open('output.txt', 'a') as f:
                f.write("\n")
                f.write('EXTMSEN-1630-Test case is passed')
                f.write("\n")

        time.sleep(5)
        # Rmx will stop the recording
        RMXstoprecording = RMXmanagerwithstoptrecording(pyautogui)
        RMXstoprecording.stoprecording()
        time.sleep(3)
        mylogin.doLoginadmin()  # login the admin portal
        time.sleep(70)  # Now it will go to Archive and check the file
        driver.find_element_by_id('a_1002').click()
        driver.find_element_by_id('a_1003').click()
        time.sleep(4)
        createdtime = driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
        print('time is', createdtime)

        try:
            myephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)  # login the EP
            myephangup.doepcallhangup()  # EP will hangup the call
        except Exception as e:
            print('EP disconnection error:', e)
            myephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)  # login the EP
            myephangup.doepcallhangup()  # EP will hangup the call
            time.sleep(2)
        if createdtime >= dt_string and pertext == 'SIP':
            print("test case is passed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=10, column=3).value = 'Pass'
            print('value is', sheet.cell(row=10, column=3).value)
            book.save('C:\\Users\\njain\\Desktop\\Report.xlsx')
        else:
            print("failed")
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=10, column=3).value = 'Fail'
            print('value is', sheet.cell(row=10, column=3).value)
            book.save('C:\\Users\\njain\\Desktop\\Report.xlsx')

        assert createdtime >= dt_string, "EXTMSEN-1630-Assert error: file is not created"
        self.assertEqual(pertext, 'SIP',"EXTMSEN-1630-Assert error: Test case is failed because protocol is not H323!")

    if __name__ == '__main__':
        unittest.main()



