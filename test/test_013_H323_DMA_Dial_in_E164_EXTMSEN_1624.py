import webbrowser
import requests
import tkinter as tk
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
import httplib2
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from Login_config.LoginHandleradminportal import Loginmsadmin
from .EPlogin_call import Eplogin
from .EPlogin_callhangup import Ephangup
import logging
from datetime import datetime
import openpyxl
import unittest
from configparser import ConfigParser



class testdialindmah323callformep(unittest.TestCase):

    def test_dial_in_dma_MS_H323_form_ep(self):

        try:
            print("Hello....")
            parser = ConfigParser()
            parser.read('config.ini')
            driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
            myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), driver)
            myepcall = Eplogin(parser.get('bug_tracker', 'epurl'),driver)  # Login the EP
            myepcall.EPH323call()
            time.sleep(4)
    
            now = datetime.now()
            dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
            type(dt_string)
            print("date and time =", dt_string)
            time.sleep(10)
            myloginadmin.doLoginadmin()
            time.sleep(40)
            pertext = driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
            print("VOIP protocol is :", pertext)
            time.sleep(5)
            driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
            time.sleep(4)
            driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()  # click OK
            time.sleep(70)  # Now it will go to Archive and check the file
            driver.find_element_by_id('a_1002').click()
            time.sleep(5)
            driver.find_element_by_id('a_1003').click()
            time.sleep(4)
            createdtime = driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
            print('time is', createdtime)
            time.sleep(2)
            if createdtime >= dt_string and pertext == 'H323':
                print("test case is passed")
                book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                sheet = book['sample']
                print('sheet is', sheet)
                sheet.cell(row=8, column=3).value = 'Pass'
                print('value is', sheet.cell(row=8, column=3).value)
                book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                print("failed")
            assert createdtime >= dt_string, "EXTMSEN-1624-Assert error: file is not created"
            self.assertEqual(pertext, 'H323',"EXTMSEN-1624-Assert error: Test case is failed because protocol is not H323!")

        finally:
            try:
                print('This is always executed')
                driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
                parser = ConfigParser()
                parser.read('config.ini')
                ephangup = Ephangup(parser.get('bug_tracker', 'epurl'),driver)
                ephangup.doepcallhangup()

            except WebDriverException as e:
                print('Result:', e)


        # driver.find_element_by_id('splitbutton-1040-btnIconEl').click()   # Hangup the call at EP

    if __name__ == '__main__':
        unittest.main()
