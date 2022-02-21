import webbrowser
import requests
import tkinter as tk
import time
import logging
from selenium import webdriver
from Login_config.LoginHandleradminportal import Loginmsadmin
from Login_config.LoginHandler import Loginms
from selenium.webdriver.support.ui import Select
from configparser import ConfigParser
import io
import unittest
from .EPlogin_call import Eplogin
from .EPlogin_callhangup import Ephangup
from datetime import datetime
from configparser import ConfigParser
import openpyxl
from selenium.common.exceptions import TimeoutException, WebDriverException


class dialindmaformep(unittest.TestCase):

    def test_dial_in_dma_MS_form_ep(self):

        try:
            print("Hello....")
            parser = ConfigParser()
            parser.read('config.ini')
            driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
            myloginadmin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), driver)
            myepcall = Eplogin(parser.get('bug_tracker', 'epurl'), driver)  # Login the EP
            myepcall.EPSIPcall()
            time.sleep(2)
            now = datetime.now()
            dt_string = now.strftime("%m-%d-%Y %I:%M%p" + ' (GMT+05:30)')
            type(dt_string)
            print("date and time =", dt_string)
            time.sleep(10)
            myloginadmin.doLoginadmin()
            time.sleep(40)
            pertext = driver.find_element_by_xpath("//*[@id='call-table-row-0']/td[6]").text
            print("VOIP protocol is :", pertext)
            logging.basicConfig(filename='output.txt', level=logging.DEBUG,
                                format='[%(levelname)s]: [%(asctime)s] [%(message)s]', datefmt='%m/%d/%Y %I:%M:%S %p')
            with open('output.txt', 'a') as f:
                f.write("\n")
            try:
                assert pertext == "SIP", "EXTMSEN-1622-Assert error: Test case is failed because protocol is not SIP!"
            except AssertionError as e:
                print('Result:', e)
                logging.error(str(e))
            if pertext == 'SIP':
                with open('output.txt', 'a') as f:
                    f.write("\n")
                    f.write('EXTMSEN-1621-Test case is passed')
                    f.write("\n")
            # driver.find_elements_by_class_name('icon-hangup')[0].click()  # Hang up the call
            time.sleep(5)
            driver.find_elements_by_class_name('portalTooltip')[0].click()  # Hang up the call.
            time.sleep(4)
            driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()  # click OK
            time.sleep(70)  # Now it will go to Archive and check the file
            driver.find_element_by_id('a_1002').click()
            time.sleep(5)
            driver.find_element_by_id('a_1003').click()
            time.sleep(4)
            createdtime = driver.find_element_by_xpath("//*[@id='archive-table-row-0']/td[4]").text
            print('time is', createdtime)
            time.sleep(2)
            if createdtime >= dt_string and pertext == 'SIP':
                print("test case is passed")
                book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
                sheet = book['sample']
                print('sheet is', sheet)
                sheet.cell(row=8, column=3).value = 'Pass'
                print('value is', sheet.cell(row=8, column=3).value)
                book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            else:
                print("failed")
            assert createdtime >= dt_string, "EXTMSEN-1621-Assert error: file is not created"
            self.assertEqual(pertext, 'SIP',"EXTMSEN-1621-Assert error: Test case is failed because protocol is not SIP!")
        finally:
            try:
                print('This is always executed')
                driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
                parser = ConfigParser()
                parser.read('config.ini')
                ephangup = Ephangup(parser.get('bug_tracker', 'epurl'), driver)
                ephangup.doepcallhangup()

            except WebDriverException as e:
                print('Result:', e)

    if __name__ == '__main__':
        unittest.main()





