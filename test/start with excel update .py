import openpyxl
from openpyxl import Workbook


def main():
    print('start')
    book = openpyxl.load_workbook('C:\\Users\\njain\\Desktop\\Report.xlsx')
    sheet = book['sample']
    print('sheet is', sheet)
    sheet.cell(row=2, column=1).value = 'EXTMSEN-1605'
    sheet.cell(row=3, column=1).value = 'EXTMSEN-1606'
    sheet.cell(row=4, column=1).value = 'EXTMSEN-1607'
    sheet.cell(row=5, column=1).value = 'EXTMSEN-1625'
    sheet.cell(row=6, column=1).value = 'EXTMSEN-1627'
    sheet.cell(row=7, column=1).value = 'EXTMSEN-1628'
    sheet.cell(row=8, column=1).value = 'EXTMSEN-1622'
    sheet.cell(row=9, column=1).value = 'EXTMSEN-1630'
    sheet.cell(row=10, column=1).value = 'EXTMSEN-1629'

    sheet.cell(row=2, column=2).value = 'SIP Registration_TCP'
    sheet.cell(row=3, column=2).value = 'SIP Registration_UDP'
    sheet.cell(row=4, column=2).value = 'SIP Registration_TLS'
    sheet.cell(row=5, column=2).value = 'P2P_Dial out_Only_H323'
    sheet.cell(row=6, column=2).value = 'RMX_Dial out_SIP'
    sheet.cell(row=7, column=2).value = 'RMX_Dial out_H323'
    sheet.cell(row=8, column=2).value = 'SIP_TLS_DMA_Dail in (sip prefix name)'
    sheet.cell(row=9, column=2).value = 'RMX_Recording Links (Dial in)'
    sheet.cell(row=10, column=2).value = 'RMX_Recording Links (Dial in)_SIP'
    for rowcolumn in range(2,20):
        sheet.cell(row=rowcolumn, column=3).value = 'Fail'
        print('value is', sheet.cell(row=rowcolumn, column=rowcolumn).value)
        book.save('C:\\Users\\njain\\Desktop\\Report.xlsx')


if __name__ == '__main__':
    main()