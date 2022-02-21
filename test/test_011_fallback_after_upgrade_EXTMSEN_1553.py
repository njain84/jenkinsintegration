import time
import logging
from selenium import webdriver
import pyautogui
from Login_config.LoginHandleradminportal import Loginmsadmin
from configparser import ConfigParser
import unittest
import openpyxl


class System_Fallback_upgrade(unittest.TestCase):

    def test_fallback_upgrade(self):
        print("Hello....")
        parser = ConfigParser()
        parser.read('config.ini')
        self.driver = webdriver.Chrome('C:/Program Files (x86)/Google/Chrome/chromedriver_win32/chromedriver.exe')
        mylogin = Loginmsadmin('admin', 123, parser.get('bug_tracker', 'url'), self.driver)
        mylogin.doLoginadmin()
        time.sleep(10)
        # click on Admin tab
        self.driver.find_element_by_xpath("//*[@id='a_1019']").click()
        time.sleep(3)
        # click on Product activation
        self.driver.find_element_by_xpath("//*[@id='a_1020']").click()
        time.sleep(3)
        version = self.driver.find_element_by_xpath("//*[@id='normalLicenseDiv']/div[2]/label[2]").text
        print('Software Version is', version)
        time.sleep(5)
        self.driver.find_element_by_id("a_1023").click()  # click fallback
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='mainForm:isAgreeCheck']").click()  #tick checkbox
        time.sleep(3)
        self.driver.find_element_by_id("fallbackButton").click()  # click on fallback Mediasuite
        time.sleep(4)
        self.driver.find_element_by_xpath("//*[@id='dialogFirstBt']").click()   # OK
        time.sleep(250)
        mylogin.doLoginadmin()
        # click on Admin tab
        time.sleep(10)
        self.driver.find_element_by_xpath("//*[@id='a_1019']").click()
        time.sleep(3)
        # click on Product activation
        self.driver.find_element_by_xpath("//*[@id='a_1020']").click()
        time.sleep(3)
        oldversion = self.driver.find_element_by_xpath("//*[@id='normalLicenseDiv']/div[2]/label[2]").text
        print('Software Version is', oldversion)
		
        time.sleep(3)
        if oldversion < version:
            book = openpyxl.load_workbook('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
            sheet = book['sample']
            print('sheet is', sheet)
            sheet.cell(row=12, column=3).value = 'Pass'
            print('value is', sheet.cell(row=12, column=3).value)
            book.save('C:\\Users\\MediaSuite\\Desktop\\Report.xlsx')
        else:
            print('Test is fail')

        assert oldversion < version, "Fallback is failed"
        time.sleep(20)

        if __name__ == '__main__':
            unittest.main()

